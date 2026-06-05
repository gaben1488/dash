"""
xlsx_style_extract.py — ВИЗУАЛЬНЫЕ метаданные листа до атомов:
объединённые диапазоны, ширины столбцов, высоты строк, и по каждой ячейке —
заливка (цвет), шрифт (bold/italic/size/цвет), границы, числовой формат,
выравнивание. Нужно, чтобы воспроизвести оформление оригинала в UI.

Usage:
  python scripts/xlsx_style_extract.py <file>                       # листы + dims + merged count
  python scripts/xlsx_style_extract.py <file> --sheet "ШДЮ старый"  # merges/widths + строки 1-12
  python scripts/xlsx_style_extract.py <file> --sheet "..." --rows 1-40
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

args = sys.argv[1:]
if not args:
    print("usage: xlsx_style_extract.py <file> [--sheet NAME] [--rows A-B]"); sys.exit(1)
file = args[0]
def opt(n): return args[args.index(n) + 1] if n in args else None
sheet = opt('--sheet'); rng_rows = opt('--rows')

wb = load_workbook(file, data_only=True)
if not sheet:
    for ws in wb.worksheets:
        print(f"  «{ws.title}»  dims={ws.dimensions}  merged={len(list(ws.merged_cells.ranges))}")
    sys.exit(0)
if sheet not in wb.sheetnames:
    print(f"НЕТ листа «{sheet}». Есть: {wb.sheetnames}"); sys.exit(1)
ws = wb[sheet]

def argb(color):
    if color is None:
        return None
    rgb = getattr(color, 'rgb', None)
    if isinstance(rgb, str) and rgb not in ('00000000', 'FFFFFFFF'):
        return rgb
    theme = getattr(color, 'theme', None)
    if theme is not None:
        tint = getattr(color, 'tint', 0) or 0
        return f"theme{theme}/{tint:+.2f}"
    return None

print(f"=== «{sheet}»  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column} ===")

merged = sorted(list(ws.merged_cells.ranges), key=lambda r: (r.min_row, r.min_col))
print(f"--- объединённые диапазоны ({len(merged)}) ---")
for mr in merged[:60]:
    print(f"  {mr}")

print("--- ширины столбцов ---")
widths = [(c, d.width) for c, d in ws.column_dimensions.items() if d.width]
for c, w in sorted(widths)[:40]:
    print(f"  {c}: {w:.1f}")

if rng_rows:
    a, b = rng_rows.split('-'); lo, hi = int(a), int(b)
else:
    lo, hi = 1, min(ws.max_row, 12)

# высоты строк в диапазоне
print(f"--- высоты строк [{lo}..{hi}] ---")
for r in range(lo, hi + 1):
    d = ws.row_dimensions.get(r)
    if d and d.height:
        print(f"  {r}: {d.height:.0f}")

print(f"--- стили ячеек [{lo}..{hi}] (только нестандартные) ---")
for r in range(lo, hi + 1):
    any_cell = False
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        fill = cell.fill
        fg = argb(fill.fgColor) if (fill and fill.patternType) else None
        font = cell.font
        bold = bool(font and font.bold)
        ital = bool(font and font.italic)
        fcolor = argb(font.color) if font else None
        size = int(font.sz) if (font and font.sz) else None
        nf = cell.number_format
        al = cell.alignment
        wrap = bool(al and al.wrap_text)
        halign = al.horizontal if al else None
        parts = []
        if fg: parts.append(f"fill={fg}")
        if bold: parts.append("B")
        if ital: parts.append("i")
        if fcolor: parts.append(f"font={fcolor}")
        if size and size not in (10, 11): parts.append(f"sz{size}")
        if nf and nf != 'General': parts.append(f"nf=[{nf}]")
        if wrap: parts.append("wrap")
        if halign: parts.append(halign)
        if parts:
            any_cell = True
            L = get_column_letter(c)
            v = cell.value
            vs = (repr(v)[:24]) if v is not None else ''
            print(f"  {L}{r}: {' '.join(parts)}  {vs}")
    if any_cell:
        print("  ·")
