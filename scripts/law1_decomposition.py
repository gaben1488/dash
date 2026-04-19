# -*- coding: utf-8 -*-
"""
law1_decomposition.py — разложить 1269 нарушений Закона 1 на причины.
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from collections import defaultdict

BASE = r'C:\Users\filat\Documents\ПЛАН-РЕЕСТР-20260419T023805Z-3-001\ПЛАН-РЕЕСТР'
DEPT_FILES = ['УО.xlsx', 'УКСиМП.xlsx', 'УД.xlsx', 'УДТХ.xlsx',
              'УЭР.xlsx', 'УИО.xlsx', 'УАГЗО.xlsx', 'УФБП.xlsx']

def num(v):
    try:
        if v is None or v == '' or v == 'Х' or v == 'х':
            return 0.0
        return float(v)
    except (ValueError, TypeError):
        return 0.0

categories = defaultdict(int)
by_grbs = defaultdict(lambda: defaultdict(int))

for f in DEPT_FILES:
    grbs = f.replace('.xlsx','')
    p = os.path.join(BASE, f)
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    sheet = 'ВСЕ' if 'ВСЕ' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        try:
            float(row[0])
        except (ValueError, TypeError):
            continue
        plan_total = num(row[10])
        fact_total = num(row[24])
        eco_total = num(row[28])
        count_eco = str(row[29] or '').strip().lower()
        dt_fact = row[16]

        if plan_total == 0:
            continue
        diff = abs(plan_total - fact_total - eco_total) / plan_total
        if diff < 0.01:
            categories['ok'] += 1
            by_grbs[grbs]['ok'] += 1
            continue

        # нарушение — категоризируем
        if fact_total == 0 and eco_total == 0:
            if dt_fact is None or str(dt_fact).strip() in ('', 'Х', 'х'):
                categories['не_завершена (нет даты факта)'] += 1
                by_grbs[grbs]['не_завершена'] += 1
            else:
                categories['есть дата факта, но суммы нули'] += 1
                by_grbs[grbs]['дата_без_сумм'] += 1
        elif count_eco in ('нет','не'):
            categories['учитывать_в_экономии=нет'] += 1
            by_grbs[grbs]['не_учитывать'] += 1
        elif fact_total > 0 and eco_total == 0 and abs(plan_total - fact_total) / plan_total < 0.01:
            categories['экономия не посчитана (fact=plan)'] += 1
            by_grbs[grbs]['без_экономии'] += 1
        elif fact_total > plan_total:
            categories['факт > плана (перерасход?)'] += 1
            by_grbs[grbs]['перерасход'] += 1
        else:
            categories['прочее (требует разбора)'] += 1
            by_grbs[grbs]['прочее'] += 1
    wb.close()

print('Декомпозиция нарушений Закона 1:')
for k, v in sorted(categories.items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')

print('\nПо ГРБС:')
for grbs in sorted(by_grbs):
    print(f'  {grbs}: {dict(by_grbs[grbs])}')
