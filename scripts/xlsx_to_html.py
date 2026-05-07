# -*- coding: utf-8 -*-
"""
xlsx_to_html.py — конвертирует CMS xlsx в standalone HTML-приложение.
Все 14 листов как навигационные табы. Цвета ячеек сохраняются.
Tabulator подключён через CDN, JSON встроен прямо в HTML.

Без зависимостей кроме openpyxl.
"""
import sys, os, json, re, html
from datetime import date, datetime

sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

P_XLSX = r'C:\Users\filat\Documents\AEMR_презентация_начальству_2026-04-20\AEMR_Консолидированная_Модель_2026-04-20.xlsx'
OUT_DIR = r'C:\Users\filat\Documents\AEMR_презентация_начальству_2026-04-20\hosted'
os.makedirs(OUT_DIR, exist_ok=True)
OUT_HTML = os.path.join(OUT_DIR, 'index.html')

print(f'[1/3] Reading {P_XLSX}...')
wb = openpyxl.load_workbook(P_XLSX, data_only=True)

ARGB_TO_CSS = {
    'FFD5E8D4': '#d5e8d4',  # green
    'FFFFF2CC': '#fff2cc',  # yellow
    'FFF4CCCC': '#f4cccc',  # red
    'FFE7E6E6': '#e7e6e6',  # grey
    'FFDDEBF7': '#ddebf7',  # blue light
    'FF1F4E79': '#1f4e79',  # blue dark (header)
}

def cell_color(cell):
    f = cell.fill
    if f and f.fgColor and f.fgColor.type == 'rgb':
        rgb = f.fgColor.rgb
        if rgb and rgb != '00000000':
            return ARGB_TO_CSS.get(rgb, '#' + rgb[2:].lower())
    return None

def cell_value(v):
    if v is None:
        return ''
    if isinstance(v, (date, datetime)):
        return v.strftime('%d.%m.%Y')
    if isinstance(v, float):
        if v.is_integer():
            return f'{int(v):,}'.replace(',', ' ')
        return f'{v:,.2f}'.replace(',', ' ')
    return str(v)

sheets_data = []
for sname in wb.sheetnames:
    ws = wb[sname]
    rows = []
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        row_cells = []
        any_value = False
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell_value(cell.value)
            if v != '':
                any_value = True
            color = cell_color(cell)
            font_bold = (cell.font and cell.font.bold) or False
            row_cells.append({
                'v': v,
                'c': color,
                'b': font_bold,
            })
        if any_value:
            rows.append(row_cells)
    sheets_data.append({
        'name': sname,
        'rows': rows,
        'col_widths': [ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width or 12
                       for c in range(1, max_col + 1)],
    })
    print(f'  [{sname}] {len(rows)} non-empty rows')

wb.close()

# Compute total stats
total_rows = sum(len(s['rows']) for s in sheets_data)
print(f'[2/3] Total: {len(sheets_data)} sheets, {total_rows} rows')

# Build HTML
print('[3/3] Building HTML...')

JSON_DATA = json.dumps(sheets_data, ensure_ascii=False)

HTML = '''<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>AEMR — Консолидированная модель учёта закупочной деятельности</title>
  <style>
    *{box-sizing:border-box}
    html,body{margin:0;padding:0;height:100%;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Calibri,sans-serif;font-size:13px;color:#1f2937;background:#f9fafb}
    .header{background:#1f4e79;color:#fff;padding:14px 22px;border-bottom:1px solid #1a3f63}
    .header h1{margin:0;font-size:18px;font-weight:600}
    .header .sub{font-size:12px;opacity:.85;margin-top:2px}
    .layout{display:flex;height:calc(100vh - 64px)}
    .sidebar{width:300px;background:#fff;border-right:1px solid #e5e7eb;overflow-y:auto;flex-shrink:0}
    .sidebar h2{margin:0;padding:12px 16px;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:#6b7280;border-bottom:1px solid #e5e7eb;background:#f9fafb}
    .sidebar ul{list-style:none;margin:0;padding:0}
    .sidebar li a{display:block;padding:10px 16px;text-decoration:none;color:#374151;font-size:13px;border-bottom:1px solid #f3f4f6;line-height:1.35}
    .sidebar li a:hover{background:#f3f4f6}
    .sidebar li a.active{background:#ddebf7;color:#1f4e79;font-weight:600;border-left:3px solid #1f4e79;padding-left:13px}
    .main{flex:1;overflow:auto;padding:18px 22px}
    .main h2{margin:0 0 4px 0;font-size:18px;color:#1f4e79}
    .meta{color:#6b7280;font-size:12px;margin-bottom:12px}
    .legend{display:flex;gap:12px;font-size:11px;margin:8px 0 14px 0;flex-wrap:wrap}
    .legend span{display:inline-flex;align-items:center;gap:6px}
    .legend .sw{width:14px;height:14px;border:1px solid #d1d5db;border-radius:2px}
    .toolbar{display:flex;gap:8px;align-items:center;margin-bottom:10px;flex-wrap:wrap}
    .toolbar input{padding:6px 10px;border:1px solid #d1d5db;border-radius:4px;font-size:12px;width:280px}
    .toolbar .badge{background:#1f4e79;color:#fff;padding:3px 8px;border-radius:10px;font-size:11px}
    table.cms{border-collapse:collapse;width:max-content;font-size:12px;background:#fff}
    table.cms th,table.cms td{border:1px solid #e5e7eb;padding:5px 8px;vertical-align:top;text-align:left;white-space:nowrap;max-width:520px;overflow:hidden;text-overflow:ellipsis}
    table.cms th{background:#1f4e79;color:#fff;font-weight:600;position:sticky;top:0;z-index:2;cursor:pointer;user-select:none}
    table.cms th:hover{background:#163a5e}
    table.cms tr:hover td{background:#f9fafb}
    table.cms td.bold{font-weight:600}
    table.cms td.num{text-align:right;font-variant-numeric:tabular-nums}
    .table-wrap{overflow:auto;max-height:calc(100vh - 220px);border:1px solid #e5e7eb;border-radius:4px;background:#fff;position:relative}
    .footer{position:fixed;bottom:6px;right:14px;font-size:10px;color:#9ca3af}
    .summary{background:#fff;border:1px solid #e5e7eb;border-radius:6px;padding:14px 18px;margin-bottom:14px}
    .summary .row{display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid #f3f4f6}
    .summary .row:last-child{border-bottom:none}
    .summary .row .k{color:#6b7280}
    .summary .row .v{font-weight:600;color:#1f2937}
    @media(max-width:768px){.sidebar{display:none}.main{padding:12px}}
  </style>
</head>
<body>
  <div class="header">
    <h1>AEMR — Консолидированная модель учёта закупочной деятельности</h1>
    <div class="sub">Версия 1.0 (черновик к согласованию) · 2026-04-20 · 14 листов · ''' + str(total_rows) + ''' строк</div>
  </div>
  <div class="layout">
    <nav class="sidebar">
      <h2>Содержание</h2>
      <ul id="nav"></ul>
    </nav>
    <main class="main">
      <h2 id="title">Загрузка…</h2>
      <div class="meta" id="meta"></div>
      <div class="legend">
        <span><span class="sw" style="background:#d5e8d4"></span>подтверждено / закон 1 ✓</span>
        <span><span class="sw" style="background:#fff2cc"></span>1 источник / non-local / разбивка не сходится</span>
        <span><span class="sw" style="background:#f4cccc"></span>закон 1 нарушен / поставщик отсутствует</span>
        <span><span class="sw" style="background:#e7e6e6"></span>export error</span>
        <span><span class="sw" style="background:#ddebf7"></span>заголовок / итог</span>
      </div>
      <div class="toolbar">
        <input type="search" id="search" placeholder="Поиск в таблице (по любой колонке)">
        <span class="badge" id="badge">0 строк</span>
      </div>
      <div class="table-wrap"><table class="cms" id="grid"><thead></thead><tbody></tbody></table></div>
    </main>
  </div>
  <div class="footer">© AEMR · ''' + datetime.now().strftime('%Y-%m-%d') + '''</div>

<script>
const DATA = ''' + JSON_DATA + ''';
let active = 0;
let sortCol = -1;
let sortDir = 1;

function escapeHtml(s){
  return String(s).replace(/[&<>"]/g, c => ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"})[c]);
}

function isNumeric(s){
  return /^-?[\\d\\s]+(?:[.,]\\d+)?$/.test(String(s).trim()) && /\\d/.test(String(s));
}

function parseNum(s){
  return parseFloat(String(s).replace(/\\s/g,'').replace(',', '.')) || 0;
}

function renderNav(){
  const ul = document.getElementById('nav');
  ul.innerHTML = DATA.map((s,i) =>
    `<li><a href="#" data-i="${i}" class="${i===active?'active':''}">${escapeHtml(s.name)} <span style="color:#9ca3af;float:right;font-size:11px">${s.rows.length}</span></a></li>`
  ).join('');
  ul.querySelectorAll('a').forEach(a => a.onclick = e => {
    e.preventDefault();
    active = +a.dataset.i;
    sortCol = -1;
    document.getElementById('search').value = '';
    renderNav();
    renderSheet();
  });
}

function renderSheet(){
  const sheet = DATA[active];
  document.getElementById('title').textContent = sheet.name;
  document.getElementById('meta').textContent = sheet.rows.length + ' строк';
  // determine header rows count: find first row that has non-bold non-headerlike cell, или просто использовать row 0 как header если все bold
  // Простая стратегия: первая строка = th, далее td.
  // Но в наших листах часто 2-3 строки заголовков. Делаем th для строк где >50% cells bold.
  const grid = document.getElementById('grid');
  const thead = grid.querySelector('thead');
  const tbody = grid.querySelector('tbody');
  thead.innerHTML = '';
  tbody.innerHTML = '';

  if (sheet.rows.length === 0){
    tbody.innerHTML = '<tr><td>Лист пуст</td></tr>';
    document.getElementById('badge').textContent = '0 строк';
    return;
  }

  // header detection: первые 1-3 строки если у них bold или color=blue
  const blueHeader = '#1f4e79';
  let headerRowCount = 0;
  for (let r = 0; r < Math.min(3, sheet.rows.length); r++){
    const row = sheet.rows[r];
    const blueOrBold = row.filter(c => c.b || c.c === blueHeader).length;
    if (blueOrBold > row.length * 0.3 && r === headerRowCount){
      headerRowCount = r + 1;
    } else if (r === 0 && row.filter(c => c.b).length === row.length){
      headerRowCount = 1;
    }
  }
  if (headerRowCount === 0) headerRowCount = 1;

  // build th from headerRowCount rows merged: take last header row as primary text
  const lastHeader = sheet.rows[headerRowCount - 1];
  const trHead = document.createElement('tr');
  lastHeader.forEach((c, i) => {
    const th = document.createElement('th');
    th.textContent = c.v || '';
    th.dataset.col = i;
    th.onclick = () => {
      if (sortCol === i) sortDir *= -1;
      else { sortCol = i; sortDir = 1; }
      renderBody();
    };
    trHead.appendChild(th);
  });
  thead.appendChild(trHead);

  renderBody();
}

function renderBody(){
  const sheet = DATA[active];
  const tbody = document.getElementById('grid').querySelector('tbody');
  const search = document.getElementById('search').value.toLowerCase();

  // skip header row(s)
  const blueHeader = '#1f4e79';
  let headerRowCount = 0;
  for (let r = 0; r < Math.min(3, sheet.rows.length); r++){
    const row = sheet.rows[r];
    const blueOrBold = row.filter(c => c.b || c.c === blueHeader).length;
    if (blueOrBold > row.length * 0.3 && r === headerRowCount){
      headerRowCount = r + 1;
    }
  }
  if (headerRowCount === 0) headerRowCount = 1;

  let dataRows = sheet.rows.slice(headerRowCount);

  // search filter
  if (search){
    dataRows = dataRows.filter(row =>
      row.some(c => String(c.v).toLowerCase().includes(search))
    );
  }

  // sort
  if (sortCol >= 0){
    dataRows = dataRows.slice().sort((a, b) => {
      const av = a[sortCol] ? a[sortCol].v : '';
      const bv = b[sortCol] ? b[sortCol].v : '';
      if (isNumeric(av) && isNumeric(bv)){
        return (parseNum(av) - parseNum(bv)) * sortDir;
      }
      return String(av).localeCompare(String(bv), 'ru') * sortDir;
    });
  }

  document.getElementById('badge').textContent = dataRows.length + ' строк' + (search ? ' (фильтр)' : '');

  // render
  const frag = document.createDocumentFragment();
  const RENDER_LIMIT = 1000;
  const limited = dataRows.slice(0, RENDER_LIMIT);
  for (const row of limited){
    const tr = document.createElement('tr');
    for (const c of row){
      const td = document.createElement('td');
      td.textContent = c.v || '';
      if (c.c) td.style.background = c.c;
      if (c.b) td.classList.add('bold');
      if (isNumeric(c.v)) td.classList.add('num');
      tr.appendChild(td);
    }
    frag.appendChild(tr);
  }
  tbody.innerHTML = '';
  tbody.appendChild(frag);
  if (dataRows.length > RENDER_LIMIT){
    const tr = document.createElement('tr');
    const td = document.createElement('td');
    td.colSpan = sheet.rows[0].length;
    td.style.background = '#fef3c7';
    td.style.textAlign = 'center';
    td.style.fontStyle = 'italic';
    td.textContent = `Показано первые ${RENDER_LIMIT} из ${dataRows.length} строк. Используйте поиск, чтобы сузить.`;
    tr.appendChild(td);
    tbody.appendChild(tr);
  }
}

document.getElementById('search').addEventListener('input', renderBody);
renderNav();
renderSheet();
</script>
</body>
</html>
'''

with open(OUT_HTML, 'w', encoding='utf-8') as f:
    f.write(HTML)

size_kb = os.path.getsize(OUT_HTML) // 1024
print(f'[OK] saved: {OUT_HTML}')
print(f'     size:  {size_kb} KB')
print(f'     sheets: {len(sheets_data)}')
print(f'     rows:   {total_rows}')
