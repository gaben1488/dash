"""
xlsx_formula_dump.py — вся математика: шаблоны формул по листам.

Фиксы v2:
  • __xludf.DUMMYFUNCTION (Google-обёртка кэша) сворачивается в один шаблон
    GS_CUSTOM — это не математика, а сохранённый результат custom-функции;
    раньше каждая ячейка казалась уникальным шаблоном (УЭР: 205 «шаблонов»).
  • детект кастомных функций игнорирует СОДЕРЖИМОЕ строковых литералов
    (раньше русские слова внутри "..." ловились как имена функций).

Usage:
  python scripts/xlsx_formula_dump.py <dir|file> [--out r.md] [--json j.json] [--max-per-sheet N]
"""
import sys, os, glob, re, json
from pathlib import Path
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

args=sys.argv[1:]
if not args:
    print("usage: xlsx_formula_dump.py <dir|file> [--out md] [--json j] [--max-per-sheet N]"); sys.exit(1)
target=args[0]
def opt(n): return args[args.index(n)+1] if n in args else None
out_path=opt('--out'); out_json=opt('--json'); max_per=int(opt('--max-per-sheet') or 120)
files=sorted(glob.glob(os.path.join(target,'**','*.xlsx'),recursive=True)) if os.path.isdir(target) else [target]
print(f"[formula_dump] {len(files)} файлов")

BUILTIN=set("""SUM SUMIF SUMIFS COUNT COUNTA COUNTIF COUNTIFS AVERAGE AVERAGEIF AVERAGEIFS IF IFS IFERROR
IFNA AND OR NOT XOR VLOOKUP HLOOKUP XLOOKUP INDEX MATCH LOOKUP TEXT VALUE CONCAT CONCATENATE LEFT RIGHT
MID LEN TRIM LOWER UPPER PROPER SUBSTITUTE REPLACE FIND SEARCH ROUND ROUNDUP ROUNDDOWN INT ABS MOD MAX MIN
MAXIFS MINIFS LARGE SMALL RANK DATE DAY MONTH YEAR TODAY NOW EOMONTH DATEDIF WEEKDAY WORKDAY NETWORKDAYS
ISBLANK ISNUMBER ISTEXT ISERROR ISNA ISLOGICAL ISDATE NA INDIRECT OFFSET ROW COLUMN ROWS COLUMNS
ARRAYFORMULA QUERY FILTER SORT UNIQUE SPLIT JOIN TEXTJOIN REGEXMATCH REGEXEXTRACT REGEXREPLACE SUMPRODUCT
TRANSPOSE CHOOSE SWITCH LET IMPORTRANGE COUNTBLANK CEILING FLOOR SIGN POWER SQRT FORMULATEXT ISEVEN ISODD
CHAR CODE TO_TEXT TO_DATE TO_PERCENT N T""".split())
REF=re.compile(r'(\$?[A-Za-z]{1,3}\$?)\d+')
STR=re.compile(r'"(?:[^"]|"")*"')
FUNC=re.compile(r'([A-Za-z_][A-Za-z0-9_\.]*)\s*\(')

def templ(f):
    if '__xludf' in f.lower(): return 'GS_CUSTOM'
    return REF.sub(r'\1#', f)
def funcs_outside_strings(f):
    bare=STR.sub('""', f)
    return set(m.upper() for m in FUNC.findall(bare))

lines=['# XLSX formula dump — математика (шаблоны формул)','']
manifest=[]
for f in files:
    name=os.path.basename(f); lines.append(f'## {name}'); fe={'file':name,'sheets':[]}
    try: wb=load_workbook(f, data_only=False, read_only=True)
    except Exception as e: lines.append(f'  ОШИБКА: {e}'); fe['error']=str(e); manifest.append(fe); continue
    for ws in wb.worksheets:
        groups=defaultdict(lambda:{'n':0,'ex':None,'raw':None}); total=0; custom=set(); gs=0
        try:
            for row in ws.iter_rows():
                for c in row:
                    v=c.value
                    if isinstance(v,str) and v.startswith('='):
                        total+=1; t=templ(v)
                        if t=='GS_CUSTOM': gs+=1
                        g=groups[t]; g['n']+=1
                        if g['ex'] is None: g['ex']=c.coordinate; g['raw']=v[:160]
                        custom|={fn for fn in funcs_outside_strings(v) if fn not in BUILTIN and not fn.startswith('__XLUDF') and fn!='DUMMYFUNCTION'}
        except Exception as e:
            lines.append(f'### «{ws.title}»: ОШИБКА {e}'); continue
        if total==0: continue
        se={'sheet':ws.title,'formula_cells':total,'templates':len(groups),'gs_custom':gs,'custom_funcs':sorted(custom)[:30],'top':[]}
        lines.append(f'### лист «{ws.title}» — {total} формул, {len(groups)} шаблонов'+(f', GS_CUSTOM×{gs}' if gs else ''))
        if custom: lines.append(f'  кастомные функции: {", ".join(sorted(custom)[:25])}')
        for t,g in sorted(groups.items(), key=lambda kv:-kv[1]['n'])[:max_per]:
            lines.append(f'  - ×{g["n"]:>4} [{g["ex"]}] `{g["raw"]}`')
            se['top'].append({'template':t[:60],'count':g['n'],'example':g['ex'],'raw':g['raw']})
        if len(groups)>max_per: lines.append(f'  … ещё {len(groups)-max_per}')
        lines.append(''); fe['sheets'].append(se)
    try: wb.close()
    except Exception: pass
    manifest.append(fe)

text='\n'.join(lines)
if out_path: Path(out_path).write_text(text,encoding='utf-8'); print(f'[OK] md {out_path} ({len(text)} chars)')
if out_json: Path(out_json).write_text(json.dumps(manifest,ensure_ascii=False,indent=1),encoding='utf-8'); print(f'[OK] json {out_json}')
if not out_path and not out_json: print(text[:6000])
