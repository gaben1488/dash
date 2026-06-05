"""
xlsx_risk_scan.py — ДЕТЕРМИНИРОВАННЫЙ движок учёта рисков по RISK_CATALOG.
Считает (не только извлекает) риски на ПРОВЕРЕННОЙ карте из DATA_TRUTH.md.

Колонко-карта листа-регистра «ВСЕ» (0-based):
  A0 id, B1 ГРБС, C2 подвед, F5 тип, G6 предмет,
  H7/I8/J9 план-источники, K10 ИТОГО-план, L11 способ, M12 причина,
  N13 план-дата, O14 кв, P15 год, Q16 факт-дата, R17 кв, S18 год, T19 откл,
  V21/W22/X23 факт-источники, Y24 ИТОГО-факт, Z25/AA26/AB27 остаток, AC28 ИТОГО-остаток,
  AD29 гейт-экономии, AE30/AF31 комментарии.
Лист-регистр детектируется по сигнатуре: header[K]~'ИТОГО' AND header[L]~'Способ' (строка 3).

Покрытые ID каталога: A-02,A-06,A-07,A-08,A-11/13(где доли), B-02(checksum),
C-01(способ),C-06(листы),C-07(ИСТОРИЯ),A-01/F-05(сдвиг D15-17), D-01/D-08(ЕП-порог, ОГОВОРКА: лимит≠НМЦК),
плюс агрегаты: доля ЕП, покрытие AD-гейта, исполнение по факт-дате, число ГРБС (A-12).
Отложено (нужны EIS-join/кросс-снимки/участники): D-02..D-13(кроме порога), E-* антикоррупция, B-01 циклы.

Usage:
  python scripts/xlsx_risk_scan.py <dir|file> [--out md] [--json j]
"""
import sys, os, glob, json, re
from pathlib import Path
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

args = sys.argv[1:]
if not args:
    print("usage: xlsx_risk_scan.py <dir|file> [--out md] [--json j]"); sys.exit(1)
target = args[0]
def opt(n): return args[args.index(n)+1] if n in args else None
out_md = opt('--out'); out_json = opt('--json')

ERRORS = {'#REF!','#DIV/0!','#VALUE!','#NAME?','#N/A','#NULL!','#NUM!','#SPILL!','#CALC!','#ERROR!'}
METHOD_OK = {'ЕП','ЭА','ЭА И АНАЛОГИ','ЭЗК','ЭК','ЭК И АНАЛОГИ','ЗАПРОС КОТИРОВОК','ОК','ОА'}
AD_OK = {'ДА','НЕТ'}
EXPECT_SVOD = {'СВОД ТД-ПМ','КОНТЕКСТ','РАСЧЕТ','ИСТОРИЯ','ОТЛАДКА','СПРАВКА'}
EP_LIMIT_TYS = 600.0     # 600 тыс.руб = п.4 ч.1 ст.93 (ОГОВОРКА: лимит, не НМЦК)
EP_BIG_TYS = 10000.0     # 10 млн.руб — крупная ЕП без конкурса
TOL = 0.01

def L(i):
    s=''; i+=1
    while i>0: i,r=divmod(i-1,26); s=chr(65+r)+s
    return s
def num(v):
    if isinstance(v,bool): return None
    if isinstance(v,(int,float)): return float(v)
    return None
def is_err(v): return isinstance(v,str) and v.strip() in ERRORS

files = sorted(glob.glob(os.path.join(target,'**','*.xlsx'),recursive=True)) if os.path.isdir(target) else [target]
print(f"[risk_scan] {len(files)} файлов")

register = []      # все находки
grbs_seen = set()
def add(sev, cid, file, sheet, msg, count=None, ratio=None, samples=None):
    register.append({'severity':sev,'catalog':cid,'file':file,'sheet':sheet,'msg':msg,
                     'count':count,'ratio':ratio,'samples':(samples or [])[:8]})

for f in files:
    fname = os.path.basename(f)
    try: wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        add('high','LOAD',fname,'-',f'не открылся: {e}'); continue
    sheetnames = wb.sheetnames
    # --- СВОД-проверки (если есть лист СВОД ТД-ПМ) ---
    if 'СВОД ТД-ПМ' in sheetnames:
        missing = EXPECT_SVOD - set(sheetnames)
        if missing: add('high','C-06',fname,'-',f'отсутствуют листы СВОД: {sorted(missing)}')
        if 'ИСТОРИЯ' in sheetnames:
            hrows = list(wb['ИСТОРИЯ'].iter_rows(values_only=True))
            snaps = max(0, sum(1 for r in hrows if any(v not in (None,'') for v in r)) - 1)
            if snaps < 3: add('high','C-07',fname,'ИСТОРИЯ',f'снимков={snaps} (<3 → EWMA/тренды невозможны)',count=snaps)
        # A-01/F-05: D15-D17 должны быть данными; если текст — сдвиг hardcode DATA_SCHEMA_
        svrows = list(wb['СВОД ТД-ПМ'].iter_rows(values_only=True))
        def svD(rn):
            idx = rn-1
            return svrows[idx][3] if idx < len(svrows) and len(svrows[idx]) > 3 else None
        cells = {f'D{r}': svD(r) for r in (14,15,16,17)}
        textish = {k:v for k,v in cells.items() if isinstance(v,str)}
        if textish:
            add('critical','A-01/F-05',fname,'СВОД ТД-ПМ',
                f'D15-D17 содержат ТЕКСТ заголовков {textish} — хардкод DATA_SCHEMA_ читал бы заголовки, не данные', samples=list(textish.keys()))
    # --- листы-регистры (по сигнатуре) ---
    for ws in wb.worksheets:
        try: rows = list(ws.iter_rows(values_only=True))
        except Exception: continue
        if len(rows) < 4: continue
        hdr = rows[2] if len(rows) > 2 else ()
        def h(i): return str(hdr[i]) if i < len(hdr) and hdr[i] else ''
        if not ('ИТОГО' in h(10).upper() and ('СПОСОБ' in h(11).upper())):
            continue  # не лист-регистр
        sheet = ws.title
        data = rows[3:]
        active = [(idx+4, r) for idx,r in enumerate(data) if r and len(r)>0 and r[0] not in (None,'')]
        n = len(active)
        if n == 0: continue
        # счётчики
        errs=[]; mixK=[]; mixY=[]; chkK=[]; chkY=[]; chkAC=[]; badMethod=[]; badAD=[]
        emptyK=[]; emptyL=[]; emptyF=[]; negK=[]; negY=[]
        ep=0; ep_plan=0.0; tot_plan=0.0; fact_filled=0; ad_set=0; econ_excluded=[]
        ep_over600=[]; ep_big=[]
        for rownum, r in active:
            def g(i): return r[i] if i < len(r) else None
            B,F,K,Lm,N,Q,Y,AC,AD = g(1),g(5),g(10),g(11),g(13),g(16),g(24),g(28),g(29)
            H,I,J,V,W,X,Z,AA,AB = g(7),g(8),g(9),g(21),g(22),g(23),g(25),g(26),g(27)
            if isinstance(B,str) and B.strip(): grbs_seen.add(B.strip())
            # errors anywhere
            for i,v in enumerate(r):
                if is_err(v): errs.append(f'{L(i)}{rownum}={v.strip()}');
            # checksum K=H+I+J
            nk,nh,ni,nj = num(K),num(H),num(I),num(J)
            if None not in (nk,nh,ni,nj):
                if abs(nk-(nh+ni+nj)) > max(TOL, abs(nk)*1e-6): chkK.append(f'K{rownum}({nk}≠{round(nh+ni+nj,3)})')
                tot_plan += nk
                if isinstance(Lm,str) and Lm.strip().upper()=='ЕП': ep_plan += nk
            elif K not in (None,'') and num(K) is None: mixK.append(f'K{rownum}={str(K)[:15]}')
            ny,nv,nw,nx = num(Y),num(V),num(W),num(X)
            if None not in (ny,nv,nw,nx):
                if abs(ny-(nv+nw+nx)) > max(TOL, abs(ny)*1e-6): chkY.append(f'Y{rownum}')
            elif Y not in (None,'') and num(Y) is None: mixY.append(f'Y{rownum}={str(Y)[:15]}')
            nac,nz,naa,nab = num(AC),num(Z),num(AA),num(AB)
            if None not in (nac,nz,naa,nab) and abs(nac-(nz+naa+nab))>max(TOL,abs(nac)*1e-6): chkAC.append(f'AC{rownum}')
            if nk is not None and nk<0: negK.append(f'K{rownum}={nk}')
            if ny is not None and ny<0: negY.append(f'Y{rownum}={ny}')
            # method
            m = Lm.strip().upper() if isinstance(Lm,str) else ''
            if not m: emptyL.append(f'L{rownum}')
            elif m not in METHOD_OK: badMethod.append(f'L{rownum}={Lm.strip()[:15]}')
            if m=='ЕП': ep+=1
            # K empty / F empty
            if K in (None,''): emptyK.append(f'K{rownum}')
            if F in (None,''): emptyF.append(f'F{rownum}')
            # fact filled (исполнение по факт-дате)
            if Q not in (None,'') and str(Q).strip().upper() not in ('Х','X'): fact_filled+=1
            # AD gate
            ad = AD.strip().upper() if isinstance(AD,str) else ''
            if ad in AD_OK: ad_set+=1
            elif AD not in (None,''): badAD.append(f'AD{rownum}={str(AD)[:10]}')
            # economy excluded: есть остаток, но AD≠да
            if nac is not None and abs(nac)>TOL and ad!='ДА': econ_excluded.append(f'AC{rownum}={round(nac,2)}')
            # 44-FZ advisory (лимит, не НМЦК)
            if m=='ЕП' and nk is not None and nk>EP_LIMIT_TYS:
                (ep_big if nk>EP_BIG_TYS else ep_over600).append(f'{rownum}:{round(nk,1)}тыс')
        # эмиссия находок
        if errs: add('critical','A-06',fname,sheet,f'ошибки формул в ячейках: {len(errs)}',count=len(errs),samples=errs)
        if chkK: add('critical','B-02',fname,sheet,f'нарушен checksum K=SUM(H:J): {len(chkK)}',count=len(chkK),samples=chkK)
        if chkY: add('high','B-02',fname,sheet,f'нарушен checksum Y=SUM(V:X): {len(chkY)}',count=len(chkY),samples=chkY)
        if chkAC: add('high','B-02',fname,sheet,f'нарушен checksum AC=SUM(Z:AB): {len(chkAC)}',count=len(chkAC),samples=chkAC)
        if mixK: add('high','A-08',fname,sheet,f'K (план) текст вместо числа: {len(mixK)}',count=len(mixK),samples=mixK)
        if mixY: add('high','A-08',fname,sheet,f'Y (факт) текст вместо числа: {len(mixY)}',count=len(mixY),samples=mixY)
        if negK: add('high','A-02',fname,sheet,f'отрицательный план K: {len(negK)}',count=len(negK),samples=negK)
        if negY: add('high','A-02',fname,sheet,f'отрицательный факт Y: {len(negY)}',count=len(negY),samples=negY)
        if badMethod: add('high','C-01',fname,sheet,f'способ вне допустимых {sorted(METHOD_OK)}: {len(badMethod)}',count=len(badMethod),samples=badMethod)
        if badAD: add('med','C-01',fname,sheet,f'AD вне {{да,нет}}: {len(badAD)}',count=len(badAD),samples=badAD)
        if emptyL: add('high','A-07',fname,sheet,f'пустой способ L: {len(emptyL)}/{n}',count=len(emptyL),ratio=round(len(emptyL)/n,3),samples=emptyL)
        if emptyK: add('high','A-07',fname,sheet,f'пустой план K: {len(emptyK)}/{n}',count=len(emptyK),ratio=round(len(emptyK)/n,3),samples=emptyK)
        ad_cov = round(ad_set/n,3)
        if ad_cov < 0.5:
            add('critical','AD-GATE',fname,sheet,f'покрытие гейта экономии AD={int(ad_cov*100)}% ({ad_set}/{n}) — экономия системно занижена',count=ad_set,ratio=ad_cov)
        if econ_excluded:
            add('high','AD-GATE',fname,sheet,f'строки с остатком, но AD≠да (экономия не учтена): {len(econ_excluded)}',count=len(econ_excluded),samples=econ_excluded)
        if ep_over600: add('med','D-01',fname,sheet,f'ЕП с лимитом >600 тыс (ОГОВОРКА: лимит≠НМЦК, нужен EIS): {len(ep_over600)}',count=len(ep_over600),samples=ep_over600)
        if ep_big: add('high','D-08',fname,sheet,f'ЕП с лимитом >10 млн (крупная без конкурса? нужен EIS): {len(ep_big)}',count=len(ep_big),samples=ep_big)
        # агрегаты (в register как info)
        ep_share_cnt = round(ep/n,3)
        ep_share_sum = round(ep_plan/tot_plan,3) if tot_plan else None
        exec_cnt = round(fact_filled/n,3)
        add('info','AGG',fname,sheet,f'активных={n} ЕП-доля(шт)={ep_share_cnt} ЕП-доля(сумма)={ep_share_sum} исполнение(факт-дата)={exec_cnt} AD-покрытие={ad_cov}',
            count=n, ratio=ep_share_cnt)
    try: wb.close()
    except Exception: pass

# A-12 число ГРБС
if grbs_seen:
    add('high' if len(grbs_seen)<8 else 'info','A-12','*','*',f'распознано ГРБС: {len(grbs_seen)} {sorted(grbs_seen)}',count=len(grbs_seen))

# --- вывод ---
sev_order = {'critical':0,'high':1,'med':2,'info':3}
register.sort(key=lambda r:(sev_order.get(r['severity'],9), r['catalog']))
by_sev = Counter(r['severity'] for r in register)
by_cat = Counter(r['catalog'] for r in register if r['severity']!='info')

md = ['# RISK REGISTER — детерминированный учёт рисков (xlsx_risk_scan)','',
      f"severity: {dict(by_sev)}",
      f"по каталогу (без info): {dict(by_cat)}",
      '',
      'Покрыто: A-01,A-02,A-06,A-07,A-08,B-02,C-01,C-06,C-07,D-01,D-08,A-12,AD-GATE.',
      'Отложено (EIS-join/кросс-снимки/участники): D-02..D-07,D-09..D-13,E-01..E-16,B-01,A-09,A-10.','']
cur=None
for r in register:
    key=(r['file'],r['sheet'])
    if key!=cur:
        cur=key; md.append(f"\n## {r['file']} :: {r['sheet']}")
    extra = (f" ratio={r['ratio']}" if r['ratio'] is not None else '') + (f" | {', '.join(r['samples'][:6])}" if r['samples'] else '')
    md.append(f"- [{r['severity']}] {r['catalog']}: {r['msg']}{extra}")
text='\n'.join(md)
if out_md: Path(out_md).write_text(text,encoding='utf-8'); print(f'[OK] md {out_md} ({len(text)} chars)')
if out_json: Path(out_json).write_text(json.dumps(register,ensure_ascii=False,indent=1),encoding='utf-8'); print(f'[OK] json {out_json}')
print(f"severity: {dict(by_sev)}")
if not out_md and not out_json: print(text[:6000])
