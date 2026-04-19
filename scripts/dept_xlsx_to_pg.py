"""Dept xlsx → PostgreSQL ingest (skeleton).

Читает один dept-xlsx (33 колонки A-AG, ~3132 строки суммарно по 8 ГРБС)
и грузит в таблицу `procurement_atoms` в Supabase/PostgreSQL.

Из memory/data/sqlite-v2.md §2.7 — целевая схема (с V2 extensions):
  - dept (canonical alias) → resolveGrbsAlias
  - method_family ('EP' | 'EA') → normalizeMethod из shared/dictionaries
  - ep_reason_canonical (15 кластеров) → canonicalizeReasonEp
  - is_org_itself (bool) — «X»/«Х»/«—» в колонке C
  - legal_ref_codes (JSON) → parseLegalRef

Skeleton только — полноценный runtime требует:
  1. Supabase self-hosted docker-compose up (postgres + supavisor + studio)
  2. Apply DDL из memory/PG_MIGRATION.md (11 таблиц)
  3. .env с SUPABASE_URL + SUPABASE_SERVICE_KEY

Usage (после настройки):
    python scripts/dept_xlsx_to_pg.py --dept УЭР --file path/to/УЭР.xlsx
    python scripts/dept_xlsx_to_pg.py --dept УО --file path/to/УО.xlsx --dry-run

Hardware-aware: читает один dept за раз, не грузит 8 параллельно.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pip install openpyxl psycopg2-binary python-dotenv", file=sys.stderr)
    sys.exit(2)

# GRBS alias map (upload from shared/dictionaries/grbs-registry.ts)
GRBS_ALIAS_MAP = {
    "УО": "УО", "УО АЕМР": "УО", "УО  АЕМР": "УО",
    "УКСиМП": "УКСиМП", "УКСиМП АЕМР": "УКСиМП",
    "УД": "УД", "УД АЕМР": "УД",
    "УДТХ": "УДТХ", "УДТХ АЕМР": "УДТХ",
    "УЭР": "УЭР", "УЭР АЕМР": "УЭР",
    "УАГЗО": "УАГЗО", "УАГиЗО": "УАГЗО", "УАГЗО АЕМР": "УАГЗО", "УАГиЗО АЕМР": "УАГЗО",
    "УФБП": "УФБП", "УФБП АЕМР": "УФБП",
    "УИО": "УИО", "УИО АЕМР": "УИО",
}

# Method family: EP = 'ЕП', EA = ['ЭА', 'ЭЗК', 'ЭК', 'ЭАС', 'ЭЕП']
METHOD_TO_FAMILY = {
    "ЕП": "EP",
    "ЭА": "EA", "ЭЗК": "EA", "ЭК": "EA", "ЭАС": "EA", "ЭЕП": "EA",
}

# Placeholders для «нет данных» в подведе (колонка C)
EMPTY_PLACEHOLDERS = {"", "X", "Х", "x", "х", "—", "–", "-", None}

# Колонки A-AG (index 0-32) — см. data/spreadsheet-anatomy.md §III
COL = {
    "N": 0, "DEPT": 1, "SUB": 2, "PROGRAM": 3, "SUBPROGRAM": 4,
    "ACTIVITY": 5, "SUBJECT": 6,
    "PLAN_FB": 7, "PLAN_KB": 8, "PLAN_MB": 9, "PLAN_TOTAL": 10,
    "METHOD": 11, "REASON_EP": 12,
    "PLAN_DATE": 13, "PLAN_Q": 14, "PLAN_Y": 15,
    "FACT_DATE": 16, "FACT_Q": 17, "FACT_Y": 18,
    "DEV_DAYS": 19, "DEV_REASON": 20,
    "FACT_FB": 21, "FACT_KB": 22, "FACT_MB": 23, "FACT_TOTAL": 24,
    "ECONOMY_FB": 25, "ECONOMY_KB": 26, "ECONOMY_MB": 27, "ECONOMY_TOTAL": 28,
    "AD_FLAG": 29,
    "COMMENT_GRBS": 30, "COMMENT_UER": 31, "COMMENT_UFBP": 32,
}


def resolve_grbs_alias(raw: str | None) -> str | None:
    """Нормализация имени ГРБС (УО АЕМР → УО, двойной пробел, опечатки)."""
    if not raw:
        return None
    cleaned = " ".join(str(raw).strip().split())  # schlop whitespace
    return GRBS_ALIAS_MAP.get(cleaned) or GRBS_ALIAS_MAP.get(cleaned.replace("и", ""))


def is_org_itself(sub: Any) -> bool:
    """C = «X» / «Х» / пусто = организация сама (_org_itself)."""
    if sub is None:
        return True
    return str(sub).strip() in {"X", "Х", "x", "х", "—", "–", "-", ""}


def normalize_method(raw: str | None) -> tuple[str | None, str | None]:
    """('ЭА' → ('ЭА', 'EA'))."""
    if not raw:
        return None, None
    method = str(raw).strip()
    return method, METHOD_TO_FAMILY.get(method)


def parse_row(row: tuple, row_idx: int, dept_canonical: str) -> dict | None:
    """Сырая строка → dict для INSERT в procurement_atoms. Возвращает None для пустой."""
    if len(row) < 7 or not row[COL["N"]]:
        return None
    if not row[COL["SUBJECT"]]:
        return None

    method_raw, method_family = normalize_method(row[COL["METHOD"]])

    sub_raw = row[COL["SUB"]]
    sub_id = "_org_itself" if is_org_itself(sub_raw) else str(sub_raw).strip()

    def safe_num(v):
        try:
            return float(v) if v else 0.0
        except (ValueError, TypeError):
            return 0.0

    return {
        "row_idx": row_idx,
        "dept_canonical": dept_canonical,
        "sub_id": sub_id,
        "sub_raw": str(sub_raw) if sub_raw else None,
        "subject": str(row[COL["SUBJECT"]])[:500],
        "program": str(row[COL["PROGRAM"]])[:200] if row[COL["PROGRAM"]] else None,
        "subprogram": str(row[COL["SUBPROGRAM"]])[:200] if row[COL["SUBPROGRAM"]] else None,
        "activity_raw": str(row[COL["ACTIVITY"]])[:100] if row[COL["ACTIVITY"]] else None,
        "method_raw": method_raw,
        "method_family": method_family,
        "reason_ep_raw": str(row[COL["REASON_EP"]])[:500] if row[COL["REASON_EP"]] else None,
        "plan_fb": safe_num(row[COL["PLAN_FB"]]),
        "plan_kb": safe_num(row[COL["PLAN_KB"]]),
        "plan_mb": safe_num(row[COL["PLAN_MB"]]),
        "plan_total": safe_num(row[COL["PLAN_TOTAL"]]),
        "plan_date": row[COL["PLAN_DATE"]],  # сохраняем исходник — parseDate на стороне PG
        "plan_q": str(row[COL["PLAN_Q"]]).strip() if row[COL["PLAN_Q"]] else None,
        "plan_y": safe_num(row[COL["PLAN_Y"]]),
        "fact_date": row[COL["FACT_DATE"]] if row[COL["FACT_DATE"]] not in EMPTY_PLACEHOLDERS else None,
        "fact_total": safe_num(row[COL["FACT_TOTAL"]]),
        "economy_total": safe_num(row[COL["ECONOMY_TOTAL"]]),
        "ad_flag": str(row[COL["AD_FLAG"]]).strip() if row[COL["AD_FLAG"]] else None,
        "comment_grbs": str(row[COL["COMMENT_GRBS"]])[:2000] if row[COL["COMMENT_GRBS"]] else None,
    }


def read_dept_xlsx(file: Path, dept_canonical: str) -> list[dict]:
    """Читает xlsx → список dict'ов для INSERT. Обрабатывает первый sheet с данными."""
    print(f"[read] {file.name} (dept={dept_canonical})")
    wb = load_workbook(file, read_only=True, data_only=True)
    # Первый sheet с данными (read_only даёт ws.max_row=None — пробуем все)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"[read] sheet='{sheet_name}'")
        atoms = []
        row_count = 0
        for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            row_count += 1
            atom = parse_row(row, i, dept_canonical)
            if atom:
                atoms.append(atom)
        print(f"[read] processed {row_count} rows, {len(atoms)} valid atoms")
        if atoms:
            wb.close()
            return atoms
    wb.close()
    return []


def insert_to_pg(atoms: list[dict], dry_run: bool = True) -> None:
    """INSERT в procurement_atoms. Requires SUPABASE_URL + SUPABASE_SERVICE_KEY в .env.

    dry_run=True — показывает sample без реальной вставки.
    """
    if dry_run:
        print(f"[insert] DRY RUN — {len(atoms)} atoms prepared, ни одной INSERT-ы")
        print(f"[insert] sample (first 2):")
        for a in atoms[:2]:
            print(f"  row_idx={a['row_idx']} subject={a['subject'][:60]!r} method={a['method_family']} plan_total={a['plan_total']:,.2f}")
        return

    # Real runtime — требует psycopg2 + valid .env
    try:
        import psycopg2
        from psycopg2.extras import execute_values
        from dotenv import load_dotenv
    except ImportError:
        print("ERROR for real run: pip install psycopg2-binary python-dotenv", file=sys.stderr)
        sys.exit(2)

    load_dotenv()
    url = os.getenv("SUPABASE_URL") or os.getenv("DATABASE_URL")
    if not url:
        print("ERROR: SUPABASE_URL или DATABASE_URL не в .env", file=sys.stderr)
        sys.exit(2)

    conn = psycopg2.connect(url)
    with conn.cursor() as cur:
        columns = list(atoms[0].keys())
        values = [tuple(a[c] for c in columns) for a in atoms]
        sql = f"INSERT INTO procurement_atoms ({', '.join(columns)}) VALUES %s"
        execute_values(cur, sql, values)
        conn.commit()
    conn.close()
    print(f"[insert] {len(atoms)} rows INSERTed")


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, type=Path, help="path to dept xlsx")
    ap.add_argument("--dept", required=True, help="dept raw name (для resolve_grbs_alias)")
    ap.add_argument("--dry-run", action="store_true", help="print sample without INSERT")
    args = ap.parse_args()

    dept_canonical = resolve_grbs_alias(args.dept)
    if not dept_canonical:
        print(f"ERROR: unknown dept '{args.dept}'. Known: {sorted(set(GRBS_ALIAS_MAP.values()))}", file=sys.stderr)
        sys.exit(2)

    if not args.file.exists():
        print(f"ERROR: {args.file} not found", file=sys.stderr)
        sys.exit(1)

    atoms = read_dept_xlsx(args.file, dept_canonical)
    print(f"[parse] {len(atoms)} atoms parsed")
    if atoms:
        print(f"[parse] method_family distribution: EP={sum(1 for a in atoms if a['method_family']=='EP')}, EA={sum(1 for a in atoms if a['method_family']=='EA')}, none={sum(1 for a in atoms if not a['method_family'])}")
    insert_to_pg(atoms, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
