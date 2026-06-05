"""
xlsx_full_extract.py — профиль КАЖДОГО столбца + риск-метрики на уровне значений.

Цель: «знать что в каждой ячейке и понимать данные» + сразу видеть риски
(каталог A: #ошибки формул, числа-как-текст, смешанный тип, отрицательные,
нарушение шкалы доли [0,1], незаполненность активных строк).

Раскладка dept/«ВСЕ» (Google export): стр.1 заголовок, 2 секции, 3 имена
столбцов, данные с 4. data_only=True → кэшированные значения (для Google
custom-функций это реальный результат; #NAME?/#REF! видны как текст).

Usage:
  python scripts/xlsx_full_extract.py <dir|file> --out profiles.md [--json out.json] [--csv-dir d] [--data-start 4]
"""
import sys, os, glob, csv, json, re, statistics
from pathlib import Path
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

args = sys.argv[1:]
if not args:
    print("usage: xlsx_full_extract.py <dir|file> --out md [--json j] [--csv-dir d] [--data-start N]"); sys.exit(1)
target = args[0]
def opt(n): return args[args.index(n)+1] if n in args else None
csv_dir = opt('--csv-dir'); out_md = opt('--out'); out_json = opt('--json')
data_start = int(opt('--data-start') or 4)
if csv_dir: os.makedirs(csv_dir, exist_ok=True)

files = sorted(glob.glob(os.path.join(target,'**','*.xlsx'), recursive=True)) if os.path.isdir(target) else [target]
print(f"[full_extract] {len(files)} файлов, data_start={data_start}")

ERRORS = {'#REF!','#DIV/0!','#VALUE!','#NAME?','#N/A','#NULL!','#NUM!','#SPILL!','#CALC!','#GETTING_DATA','#ERROR!'}
SHARE_RE = re.compile(r'(?i)(дол[яи]|%|процент|share|pct|удельн)')
NUMTEXT_RE = re.compile(r'^-?[\d\s  ]+[.,]?\d*%?$')

def col_letter(n):
    s=''
    while n>0:
        n,r=divmod(n-1,26); s=chr(65+r)+s
    return s

def as_ru_number(s):
    t = str(s).replace('\xa0','').replace(' ','').replace(' ','').replace('%','').replace(',','.')
    try: return float(t)
    except Exception: return None

manifest = []
prof = ['# XLSX full data extract — профиль столбцов + риск-метрики значений', '']
for f in files:
    name = os.path.splitext(os.path.basename(f))[0]
    prof.append(f'## {name}.xlsx')
    fe = {'file': os.path.basename(f), 'sheets': []}
    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        prof.append(f'  ОШИБКА: {e}'); fe['error']=str(e); manifest.append(fe); continue
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        nrows = len(rows); ncols = max((len(r) for r in rows), default=0)
        if csv_dir:
            safe = f"{name}__{ws.title}".replace('/','_').replace('\\','_')[:120]
            try:
                with open(os.path.join(csv_dir,safe+'.csv'),'w',newline='',encoding='utf-8') as cf:
                    w=csv.writer(cf)
                    for r in rows: w.writerow(['' if v is None else v for v in r])
            except Exception as e:
                prof.append(f'  [csv fail {ws.title}: {e}]')
        hdr_idx = data_start-2 if nrows>=data_start-1 else 0
        header = rows[hdr_idx] if 0<=hdr_idx<nrows else ()
        data = rows[data_start-1:] if nrows>=data_start else []
        active = [r for r in data if r and len(r)>0 and r[0] not in (None,'')]
        nactive = len(active)
        nonempty_rows = sum(1 for r in data if any(v not in (None,'') for v in r))
        se = {'sheet':ws.title,'nrows':nrows,'ncols':ncols,'data_rows':nonempty_rows,'active_rows':nactive,'columns':[],'risks':[]}
        prof.append(f'### лист «{ws.title}» — {nrows}×{ncols}, данных≈{nonempty_rows}, активных(A≠∅)={nactive}')
        for ci in range(min(ncols,45)):
            vals = [r[ci] for r in data if ci<len(r) and r[ci] not in (None,'')]
            if not vals: continue
            h = header[ci] if ci<len(header) and header[ci] not in (None,'') else '—'
            letter = col_letter(ci+1)
            cnt = Counter(str(v)[:40] for v in vals); distinct = len(cnt)
            numeric = [v for v in vals if isinstance(v,(int,float)) and not isinstance(v,bool)]
            nnum = len(numeric)
            errs = Counter(v.strip() for v in vals if isinstance(v,str) and v.strip() in ERRORS)
            nerr = sum(errs.values())
            numtext = sum(1 for v in vals if isinstance(v,str) and v.strip() not in ERRORS and as_ru_number(v) is not None and NUMTEXT_RE.match((v or '').strip()))
            kind = 'число' if nnum>len(vals)*0.7 else ('текст' if nnum==0 else 'смеш')
            filled_active = sum(1 for r in active if ci<len(r) and r[ci] not in (None,''))
            blank_ratio = round(1-filled_active/nactive,3) if nactive else None
            col = {'letter':letter,'header':str(h)[:60],'fill':len(vals),'distinct':distinct,'kind':kind,
                   'numeric':nnum,'errors':dict(errs),'numbers_as_text':numtext,'blank_ratio_active':blank_ratio}
            tags=[]
            if numeric:
                mn=min(numeric); mx=max(numeric); neg=sum(1 for v in numeric if v<0); zeros=sum(1 for v in numeric if v==0)
                col.update({'min':mn,'max':mx,'neg':neg,'zeros':zeros})
                if neg>0: tags.append(f'neg={neg}')
                if SHARE_RE.search(str(h)):
                    med=statistics.median(numeric); col['median']=round(med,4)
                    if mx>1.0 or mn<0: tags.append(f'scale_out_of_unit(max={round(mx,3)})')
                    if med>1.5: tags.append(f'scale_0_100?(median={round(med,3)})')
            if nerr>0: tags.append(f'ERR={nerr}{dict(errs)}')
            if kind=='смеш' and nnum>0: tags.append('mixed_type')
            if numtext>0: tags.append(f'num_as_text={numtext}')
            if blank_ratio is not None and blank_ratio>0.05 and nactive>=10: tags.append(f'blank={int(blank_ratio*100)}%')
            top='; '.join(f'{k}×{n}' for k,n in cnt.most_common(5))
            if tags:
                col['risk']=tags
                se['risks'].append({'col':letter,'header':str(h)[:40],'flags':tags})
            prof.append(f'- [{letter}] «{str(h)[:42]}»: запол={len(vals)} уник={distinct} тип={kind}'
                        + (f' min={col.get("min")} max={col.get("max")}' if numeric else '')
                        + (f' ⚠{",".join(tags)}' if tags else '')
                        + f' | топ: {top[:140]}')
            se['columns'].append(col)
        fe['sheets'].append(se); prof.append('')
    try: wb.close()
    except Exception: pass
    manifest.append(fe)

text='\n'.join(prof)
if out_md: Path(out_md).write_text(text,encoding='utf-8'); print(f'[OK] md {out_md} ({len(text)} chars)')
if out_json: Path(out_json).write_text(json.dumps(manifest,ensure_ascii=False,indent=1),encoding='utf-8'); print(f'[OK] json {out_json}')
if not out_md and not out_json: print(text[:6000])
