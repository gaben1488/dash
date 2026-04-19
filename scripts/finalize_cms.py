# -*- coding: utf-8 -*-
"""
finalize_cms.py — добавляет 14_ЗАКОН_1_ДЕКОМПОЗИЦИЯ и обновляет README с финальными цифрами.
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

P_CMS = r'C:\Users\filat\Documents\AEMR_презентация_начальству_2026-04-20\AEMR_Консолидированная_Модель_2026-04-20.xlsx'
BORDER = Border(*[Side(style='thin', color='FFBFBFBF')]*4)
C_HEADER = 'FF1F4E79'
FILL_HEADER = PatternFill('solid', fgColor=C_HEADER)
FILL_SUB = PatternFill('solid', fgColor='FFDDEBF7')
FILL_GREEN = PatternFill('solid', fgColor='FFD5E8D4')
FILL_YELLOW = PatternFill('solid', fgColor='FFFFF2CC')
FILL_RED = PatternFill('solid', fgColor='FFF4CCCC')

wb = openpyxl.load_workbook(P_CMS)

# Добавляем лист 14_ЗАКОН_1_ДЕКОМПОЗИЦИЯ
if '14_ЗАКОН_1_ДЕКОМПОЗ' not in wb.sheetnames:
    ws = wb.create_sheet('14_ЗАКОН_1_ДЕКОМПОЗ')
else:
    ws = wb['14_ЗАКОН_1_ДЕКОМПОЗ']
    # clear
    for r in range(1, 200):
        for c in range(1, 15):
            ws.cell(row=r, column=c).value = None

ws.cell(row=1, column=1, value='Декомпозиция нарушений Закона 1 (ПЛАН = ФАКТ + ЭКОНОМИЯ)').font = Font(size=14, bold=True, color='FF1F4E79')
ws.cell(row=2, column=1, value='Всего строк Tier 1: 3 132. Закон 1 соблюдается: 1 862. Нарушений: 1 269 (40,5%).').font = Font(italic=True, color='FF808080')
ws.cell(row=3, column=1, value='Причина 40,5% нарушений — не «ошибки в данных», а незавершённость процедур на момент выгрузки.').font = Font(italic=True, color='FF808080')

# Общая таблица
headers = ['Категория', 'Строк', 'Доля от всего', 'Это ошибка?', 'Почему так']
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFFFF')
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = BORDER

data = [
    ('Соблюдается (план=факт+экономия)', 1862, 0.5945, 'Нет',
     'Правильно заполненные процедуры. 60% реестра — достоверны.'),
    ('УО: дата факта стоит, но суммы факта/экономии = 0', 806, 0.2572, 'Нет, но требует дозаполнения',
     'Специфика УО: дата подписания контракта проставляется раньше, чем вносятся конкретные суммы из контракта. Это не ошибка, а разрыв во времени между правкой даты и правкой сумм.'),
    ('Не завершена (нет даты факта)', 414, 0.1322, 'Нет',
     'Процедура запланирована, но ещё не прошла. Ожидаемо на момент выгрузки в начале финансового года.'),
    ('Прочее (требует индивидуального разбора)', 49, 0.0156, 'Возможно',
     'Малочисленные случаи: факт > плана (перерасход), необъяснимые расхождения. Ручная проверка рекомендуется.'),
]
r = 6
for row in data:
    for ci, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=v)
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if ci == 3:
            c.number_format = '0.00%'
    # colouring
    if row[3] == 'Нет':
        ws.cell(row=r, column=4).fill = FILL_GREEN
    elif 'Возможно' in row[3]:
        ws.cell(row=r, column=4).fill = FILL_YELLOW
    else:
        ws.cell(row=r, column=4).fill = FILL_RED
    r += 1

# Итого
ws.cell(row=r, column=1, value='ИТОГО').font = Font(bold=True)
ws.cell(row=r, column=2, value=f'=SUM(B6:B{r-1})').font = Font(bold=True)
ws.cell(row=r, column=2).number_format = '#,##0'
ws.cell(row=r, column=3, value=f'=SUM(C6:C{r-1})').font = Font(bold=True)
ws.cell(row=r, column=3).number_format = '0.00%'
for ci in range(1, 6):
    ws.cell(row=r, column=ci).fill = FILL_SUB

# Декомпозиция по ГРБС
r += 3
ws.cell(row=r, column=1, value='Разрез по ГРБС').font = Font(size=12, bold=True, color='FF1F4E79')
r += 1
hdr2 = ['ГРБС', 'ok (закон 1 ✓)', 'Дата без сумм', 'Не завершена', 'Прочее', 'Всего', '% соблюдения']
for i, h in enumerate(hdr2, start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFFFF')
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = BORDER
r += 1

grbs_breakdown = [
    ('УО',     1317, 806, 5,   0, 2128),
    ('УКСиМП', 250,  0,   235, 44, 529),
    ('УД',     134,  0,   29,  0, 163),
    ('УЭР',    37,   0,   49,  0, 86),
    ('УДТХ',   45,   0,   21,  0, 66),
    ('УИО',    28,   0,   34,  0, 62),
    ('УАГЗО',  31,   0,   16,  0, 47),
    ('УФБП',   20,   0,   25,  5, 50),
]
for row in grbs_breakdown:
    grbs, ok, d_ws, ne, ot, tot = row
    pct = ok / tot if tot else 0
    for ci, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=v)
        c.border = BORDER
        c.alignment = Alignment(horizontal='center')
        if ci == 1:
            c.font = Font(bold=True)
    c_pct = ws.cell(row=r, column=7, value=pct)
    c_pct.number_format = '0.00%'
    c_pct.border = BORDER
    if pct >= 0.85:
        c_pct.fill = FILL_GREEN
    elif pct >= 0.50:
        c_pct.fill = FILL_YELLOW
    else:
        c_pct.fill = FILL_RED
    r += 1

# Выводы для начальства
r += 2
ws.cell(row=r, column=1, value='Выводы для начальства').font = Font(size=12, bold=True, color='FF1F4E79')
r += 1
conclusions = [
    '1. Данные в исходных 8 план-реестрах ГРБС в целом честные: 60% строк арифметически сходятся, 40% — объяснимы (незавершённые процедуры или разрыв во времени между проставлением даты и сумм).',
    '2. Специфика УО (806 строк с датой, но без сумм) — это не ошибка ввода, а операционный паттерн. Нужно либо дисциплинированно заполнять суммы одновременно с датой, либо в CMS считать такие строки как «в работе» и не применять к ним Закон 1.',
    '3. УКСиМП и УИО имеют высокий % незавершённых (44% и 55% соответственно). На момент выгрузки реестра это нормально, но к концу года должно сойтись ≥85%.',
    '4. УО, наоборот, даёт 61,9% соблюдения Закона 1 — самый высокий абсолютный объём корректных данных (1317 строк).',
    '5. Только 49 строк требуют ручного разбора (1,5%) — посильная нагрузка для одного куратора УФБП за одну рабочую смену.',
]
for c_text in conclusions:
    ws.cell(row=r, column=1, value=c_text).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 40
    r += 1

# autosize
widths = [42, 12, 14, 14, 12, 10, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# обновить README
ws0 = wb['00_README']
# найти последнюю строку
r = ws0.max_row + 2
ws0.cell(row=r, column=1, value='ИТОГ ЗАЛИВКИ (21 лист)').font = Font(bold=True, size=12, color='FF1F4E79')
r += 1
summary = [
    ('Всего листов в книге', 14),
    ('Tier 1 атомов (из 8 dept-xlsx)', '3 132'),
    ('Tier 5 строк (еженедельный)', '4 708'),
    ('Tier 6 строк (ежедневный)', '252'),
    ('Закон 1 соблюдается', '1 862 строки (59,5%)'),
    ('Закон 1 объяснимо нарушен', '1 220 строк (незавершённые + UO дата-без-сумм)'),
    ('Закон 1 требует разбора', '49 строк (1,5%)'),
    ('Закон 2 соблюдается', '3 131 из 3 132 (99,97%)'),
    ('ER auto-match T1↔T5', '81 пара (0,04% — подтверждает необходимость ER-слоя)'),
    ('ER auto-match T1↔T6', '75 пар'),
    ('Тройных подтверждений T1+T5+T6', '22 строки (наивысшая достоверность)'),
]
for k, v in summary:
    ws0.cell(row=r, column=1, value=k)
    ws0.cell(row=r, column=2, value=v)
    r += 1

wb.save(P_CMS)
print(f'[OK] finalized: {P_CMS}')
print(f'     sheets:  {len(wb.sheetnames)}')
print(f'     size:    {os.path.getsize(P_CMS)} bytes')
