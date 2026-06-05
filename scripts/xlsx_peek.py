"""
xlsx_peek.py — точечный дамп ячеек: ЗНАЧЕНИЕ (data_only) + ФОРМУЛА рядом.
Истина — в ячейках, не в документации. Инструмент верификации маппинга/рисков.

Usage:
  python scripts/xlsx_peek.py <file>                          # список листов + dims
  python scripts/xlsx_peek.py <file> --sheet "ВСЕ" --rows 1-6 # строки 1..6 всех столбцов
  python scripts/xlsx_peek.py <file> --sheet "СВОД ТД-ПМ" --range A12:K18
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

args = sys.argv[1:]
if not args:
    print("usage: xlsx_peek.py <file> [--sheet NAME] [--rows A-B | --range A1:K9]"); sys.exit(1)
file = args[0]
def opt(n): return args[args.index(n)+1] if n in args else None
sheet = opt('--sheet'); rng = opt('--range'); rows = opt('--rows')

wv = load_workbook(file, data_only=True)
wf = load_workbook(file, data_only=False)
if not sheet:
    print(f"FILE: {file}")
    for ws in wv.worksheets:
        print(f'  «{ws.title}»  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column}')
    sys.exit(0)
if sheet not in wv.sheetnames:
    print(f"НЕТ листа «{sheet}». Есть: {wv.sheetnames}"); sys.exit(1)
wsv = wv[sheet]; wsf = wf[sheet]
if rng:
    minc, minr, maxc, maxr = range_boundaries(rng)
elif rows:
    a, b = rows.split('-'); minr, maxr = int(a), int(b); minc, maxc = 1, wsv.max_column
else:
    minr, maxr = 1, 6; minc, maxc = 1, wsv.max_column
print(f"«{sheet}» [{minr}..{maxr}] × cols[{get_column_letter(minc)}..{get_column_letter(maxc)}]")
for r in range(minr, maxr + 1):
    any_cell = False
    for c in range(minc, maxc + 1):
        v = wsv.cell(r, c).value
        fc = wsf.cell(r, c).value
        if v is None and fc is None:
            continue
        any_cell = True
        L = get_column_letter(c)
        f = f'  FORMULA={fc}' if isinstance(fc, str) and fc.startswith('=') else ''
        print(f'  {L}{r}: {repr(v)}{f}')
    if any_cell:
        print('  ---')
