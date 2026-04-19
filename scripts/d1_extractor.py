"""D1 mega-extractor — читает все 9 исходных xlsx и пишет JSON-артефакты
для агентов α/β/γ swarm'а Discovery.
"""
import openpyxl, json, os, re, sys
from pathlib import Path
from collections import Counter
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
SRC = Path(r'C:/Users/filat/Documents/ПЛАН-РЕЕСТР-20260417T151357Z-3-001/ПЛАН-РЕЕСТР')
OUT = Path(r'C:/Users/filat/.claude/projects/C--Users-filat-dash/memory/artifacts/d1/raw')
OUT.mkdir(parents=True, exist_ok=True)


def ser(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


# === ALPHA: global своды ===
alpha = {'SVOD_GOOGLE': {}, 'SVOD_25_26': {}}

for fname, key in [('СВОД_ДЛЯ_GOOGLE.xlsx', 'SVOD_GOOGLE'),
                   ('СВОД -25-26.xlsx', 'SVOD_25_26')]:
    path = SRC / fname
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= 6:
                break
            rows.append([ser(c) for c in r[:15]])
        alpha[key][sn] = {'rows_preview_6': rows}
    wb.close()

# SVOD special sheets full
wb = openpyxl.load_workbook(str(SRC / 'СВОД_ДЛЯ_GOOGLE.xlsx'), data_only=True, read_only=True)
for sname in ['ИСТОРИЯ', 'СПРАВКА', 'КОНТЕКСТ', 'РАСЧЕТ', 'Settings', 'Контроль']:
    if sname in wb.sheetnames:
        ws = wb[sname]
        alpha[f'SVOD_full_{sname}'] = [[ser(c) for c in r]
                                       for r in ws.iter_rows(values_only=True)]
wb.close()

# ГРБС registry
wb = openpyxl.load_workbook(str(SRC / 'СВОД -25-26.xlsx'), data_only=True, read_only=True)
if 'ГРБС' in wb.sheetnames:
    ws = wb['ГРБС']
    alpha['GRBS_registry'] = [[ser(c) for c in r]
                               for r in ws.iter_rows(values_only=True)]
wb.close()

# Mirror sync — 10 rows each dept
alpha['mirror_sync'] = {}
PRIMARY = {'УО': ('УО.xlsx', 'ВСЕ'), 'УИО': ('УИО.xlsx', 'УИО'),
           'УЭР': ('УЭР.xlsx', 'ВСЕ'), 'УАГЗО': ('УАГЗО.xlsx', 'ВСЕ'),
           'УФБП': ('УФБП.xlsx', 'УФБП'), 'УД': ('УД.xlsx', 'ВСЕ'),
           'УДТХ': ('УДТХ.xlsx', 'УДТХ'), 'УКСиМП': ('УКСиМП.xlsx', 'ВСЕ')}

for dept, (pfile, psheet) in PRIMARY.items():
    try:
        wbp = openpyxl.load_workbook(str(SRC / pfile), data_only=True, read_only=True)
        if psheet not in wbp.sheetnames:
            wbp.close()
            continue
        wsp = wbp[psheet]
        primary_rows = []
        for i, r in enumerate(wsp.iter_rows(min_row=4, values_only=True)):
            if i >= 10:
                break
            if r[0] is None:
                continue
            primary_rows.append([ser(c) for c in r[:15]])
        wbp.close()
        wbm = openpyxl.load_workbook(str(SRC / 'СВОД_ДЛЯ_GOOGLE.xlsx'),
                                      data_only=True, read_only=True)
        if dept in wbm.sheetnames:
            wsm = wbm[dept]
            mirror_rows = []
            for i, r in enumerate(wsm.iter_rows(min_row=4, values_only=True)):
                if i >= 10:
                    break
                if r[0] is None:
                    continue
                mirror_rows.append([ser(c) for c in r[:15]])
            alpha['mirror_sync'][dept] = {'primary': primary_rows,
                                           'mirror': mirror_rows,
                                           'match': primary_rows == mirror_rows}
        wbm.close()
    except Exception as e:
        alpha['mirror_sync'][dept] = {'error': str(e)}

(OUT / 'alpha.json').write_text(json.dumps(alpha, ensure_ascii=False,
                                            indent=2, default=str),
                                 encoding='utf-8')
print('alpha.json OK')


# === BETA: 4 больших ===
beta = {}
BIG_FILES = [('УО.xlsx', 'ВСЕ'), ('УКСиМП.xlsx', 'ВСЕ'),
             ('УД.xlsx', 'ВСЕ'), ('УДТХ.xlsx', 'УДТХ')]

for fname, sheet in BIG_FILES:
    dept = fname.replace('.xlsx', '')
    wb = openpyxl.load_workbook(str(SRC / fname), data_only=True, read_only=True)
    beta[dept] = {}

    # ChangeLog
    if '_ChangeLog' in wb.sheetnames:
        ws = wb['_ChangeLog']
        rows = list(ws.iter_rows(values_only=True))
        headers = [ser(c) for c in rows[0]] if rows else []
        data = rows[1:]
        authors = Counter()
        priorities = Counter()
        statuses = Counter()
        cols_touched = Counter()
        pairs = []
        dmin = dmax = None
        for r in data:
            if not r or len(r) < 10:
                continue
            sheet_c, cell, col, row_n, before, after, ts, author, prio, status = r[:10]
            if author:
                authors[str(author)] += 1
            if prio:
                priorities[str(prio)] += 1
            if status:
                statuses[str(status)] += 1
            if col is not None:
                cols_touched[str(col)] += 1
            if ts:
                t = ts.isoformat() if isinstance(ts, datetime) else str(ts)
                if dmin is None or t < dmin:
                    dmin = t
                if dmax is None or t > dmax:
                    dmax = t
            if len(pairs) < 30 and before is not None and after is not None:
                pairs.append({'was': str(before)[:60], 'became': str(after)[:60]})
        beta[dept]['changelog'] = {
            'rows': len(data), 'headers': headers,
            'date_range': [dmin, dmax],
            'top_authors': dict(authors.most_common(10)),
            'priorities': dict(priorities),
            'statuses': dict(statuses),
            'top_columns_touched': dict(cols_touched.most_common(10)),
            'sample_before_after': pairs,
        }

    # Settings + Контроль
    for s in ['Settings', 'Контроль']:
        if s in wb.sheetnames:
            ws = wb[s]
            beta[dept][s.lower()] = [[ser(c) for c in r]
                                      for r in ws.iter_rows(values_only=True)]

    # Subordinates
    service = {'ВСЕ', '_ChangeLog', 'Settings', 'GOOGLE_ФОРМУЛЫ',
               'Контроль', dept}
    subords = {}
    for sn in wb.sheetnames:
        if sn in service:
            continue
        ws = wb[sn]
        total_limit = 0
        rows_count = 0
        sub_name = None
        for r in ws.iter_rows(min_row=4, values_only=True):
            if not r or r[0] is None:
                continue
            rows_count += 1
            if sub_name is None and r[2]:
                sub_name = str(r[2])[:60]
            try:
                if r[10] is not None:
                    total_limit += float(r[10])
            except (ValueError, TypeError):
                pass
        subords[sn] = {'sub_name': sub_name, 'rows': rows_count,
                       'total_limit_K': round(total_limit, 1)}
    beta[dept]['subordinates'] = subords

    # AE analysis on main sheet
    ws = wb[sheet]
    ae_stats = Counter()
    ae_mentions = {'split': 0, 'transfer': 0, 'no_financing': 0,
                   'supplier_refused': 0, 'decree_112': 0}
    re_split = re.compile(r'(?i)(разбит|раздел|част|нескольк|добавлена позиция)')
    re_transfer = re.compile(r'(?i)(переносит|перенос|перенесен)')
    re_nofin = re.compile(r'(?i)(отсутств.*финанс|нет финанс)')
    re_refused = re.compile(r'(?i)(отказыва.*поставщик|не да.*КП)')
    re_decree = re.compile(r'(?i)(распоряж.*112|112.*03\.09\.2025)')
    status_markers = Counter()
    data_rows = 0
    ae_fill = 0
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or len(r) < 33:
            continue
        if r[0] is None and r[1] is None:
            continue
        data_rows += 1
        ae = r[30]
        if ae is None:
            continue
        ae_s = str(ae).strip()
        if not ae_s:
            continue
        ae_fill += 1
        low = ae_s.lower()
        if 'договор заключен' in low or 'контракт заключен' in low:
            status_markers['signed'] += 1
        elif 'планирование' in low:
            status_markers['planning'] += 1
        elif 'аукцион не состоял' in low:
            status_markers['auction_failed'] += 1
        elif 'возвращ' in low:
            status_markers['returned'] += 1
        if re_split.search(ae_s):
            ae_mentions['split'] += 1
        if re_transfer.search(ae_s):
            ae_mentions['transfer'] += 1
        if re_nofin.search(ae_s):
            ae_mentions['no_financing'] += 1
        if re_refused.search(ae_s):
            ae_mentions['supplier_refused'] += 1
        if re_decree.search(ae_s):
            ae_mentions['decree_112'] += 1
        ae_stats[ae_s[:80]] += 1
    beta[dept]['ae_stats'] = {
        'data_rows': data_rows, 'ae_filled': ae_fill,
        'ae_fill_pct': round(100 * ae_fill / max(1, data_rows), 1),
        'top_values': dict(ae_stats.most_common(20)),
        'status_markers': dict(status_markers),
        'pattern_mentions': ae_mentions,
    }

    # Date mismatch
    pd_fd = 0
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or len(r) < 33:
            continue
        pd, fd = r[13], r[16]
        if pd and fd and pd != fd:
            pd_fd += 1
    beta[dept]['date_mismatch'] = pd_fd

    wb.close()

(OUT / 'beta.json').write_text(json.dumps(beta, ensure_ascii=False,
                                           indent=2, default=str),
                                encoding='utf-8')
print('beta.json OK')


# УО special sheets
uo_specials = {}
wb = openpyxl.load_workbook(str(SRC / 'УО.xlsx'), data_only=True, read_only=True)
for sn in ['Лист5', 'Лист6', 'Лист7', 'Лист8', 'Лист9', 'Совместные закупки']:
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    sub_names = []
    rows_count = 0
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or r[0] is None:
            continue
        rows_count += 1
        if r[2] and len(sub_names) < 5:
            sub_names.append(str(r[2])[:80])
    uo_specials[sn] = {'rows': rows_count, 'sub_names': sub_names}
wb.close()
(OUT / 'uo_special.json').write_text(json.dumps(uo_specials,
                                                 ensure_ascii=False, indent=2),
                                      encoding='utf-8')
print('uo_special.json OK')


# === GAMMA: 4 маленьких ===
gamma = {}
SMALL_FILES = [('УАГЗО.xlsx', 'ВСЕ'), ('УИО.xlsx', 'УИО'),
               ('УФБП.xlsx', 'УФБП'), ('УЭР.xlsx', 'ВСЕ')]

for fname, sheet in SMALL_FILES:
    dept = fname.replace('.xlsx', '')
    wb = openpyxl.load_workbook(str(SRC / fname), data_only=True, read_only=True)
    gamma[dept] = {'sheetnames': wb.sheetnames}

    if sheet in wb.sheetnames:
        ws = wb[sheet]
        ae_all = []
        ad_counts = Counter()
        af_prefix = Counter()
        data_rows = 0
        for r in ws.iter_rows(min_row=4, values_only=True):
            if not r or len(r) < 33:
                continue
            if r[0] is None and r[1] is None:
                continue
            data_rows += 1
            ae = r[30]
            if ae and str(ae).strip():
                ae_all.append(str(ae)[:150])
            ad = r[29]
            ad_counts[str(ad).strip() if ad else '<EMPTY>'] += 1
            af = r[31]
            if af:
                m = re.match(r'([А-ЯЁA-Z]{2,4})', str(af).strip())
                if m:
                    af_prefix[m.group(1)] += 1
        gamma[dept]['data_rows'] = data_rows
        gamma[dept]['ae_filled'] = len(ae_all)
        gamma[dept]['ae_all'] = ae_all
        gamma[dept]['ad_distribution'] = dict(ad_counts)
        gamma[dept]['af_prefixes'] = dict(af_prefix.most_common())

    if 'Контроль' in wb.sheetnames:
        ws = wb['Контроль']
        gamma[dept]['kontrol'] = [[ser(c) for c in r]
                                    for r in ws.iter_rows(values_only=True)]
    if 'Settings' in wb.sheetnames:
        ws = wb['Settings']
        gamma[dept]['settings'] = [[ser(c) for c in r]
                                     for r in ws.iter_rows(values_only=True)]

    if dept in ('УАГЗО', 'УЭР'):
        service = {'ВСЕ', '_ChangeLog', 'Settings', 'GOOGLE_ФОРМУЛЫ',
                   'Контроль', dept, 'Лист1'}
        subords = {}
        for sn in wb.sheetnames:
            if sn in service:
                continue
            ws_sub = wb[sn]
            rows_count = 0
            sub_name = None
            for r in ws_sub.iter_rows(min_row=4, values_only=True):
                if not r or r[0] is None:
                    continue
                rows_count += 1
                if sub_name is None and r[2]:
                    sub_name = str(r[2])[:60]
            subords[sn] = {'sub_name': sub_name, 'rows': rows_count}
        gamma[dept]['subordinates'] = subords

    wb.close()

(OUT / 'gamma.json').write_text(json.dumps(gamma, ensure_ascii=False,
                                            indent=2, default=str),
                                 encoding='utf-8')
print('gamma.json OK')

for f in sorted(OUT.glob('*.json')):
    print(f'  {f.name}: {os.path.getsize(f)//1024}KB')
