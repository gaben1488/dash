"""
xlsx_metadata_map.py — метаданные + риск-сигналы контроля листа.

Достаёт слой, невидимый в значениях: data validation (допустимые значения),
условное форматирование (визуальные пороги), защита, именованные диапазоны,
объединённые/скрытые ячейки, комментарии, внешние ссылки, автофильтр,
числовые форматы (%/₽), и АНОМАЛИИ КОНСИСТЕНТНОСТИ ФОРМУЛ по столбцу
(ячейка с формулой, отличной от эталона столбца = ручная подмена/инъекция;
каталог B/A). Google custom-функции (__xludf) сворачиваются в GS_CUSTOM,
иначе каждая ячейка ложно «уникальна».

Usage:
  python scripts/xlsx_metadata_map.py <dir|file> [--out r.md] [--json j.json] [--validations-only] [--expect "S1,S2"]
"""
import sys, os, glob, json, re
from pathlib import Path
from collections import Counter, defaultdict

sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

args=sys.argv[1:]
if not args:
    print("usage: xlsx_metadata_map.py <dir|file> [--out md] [--json j] [--validations-only] [--expect S1,S2]"); sys.exit(1)
target=args[0]
def opt(n): return args[args.index(n)+1] if n in args else None
out_path=opt('--out'); out_json=opt('--json'); val_only='--validations-only' in args
expect=set(s.strip() for s in (opt('--expect') or '').split(',') if s.strip())

files = sorted(glob.glob(os.path.join(target,'**','*.xlsx'),recursive=True)) if os.path.isdir(target) else ([target] if target.endswith('.xlsx') else [])
print(f"[xlsx_metadata_map] {len(files)} файлов")

REF=re.compile(r'(\$?[A-Za-z]{1,3}\$?)\d+')
def templ(fml):
    if '__xludf' in fml.lower(): return 'GS_CUSTOM'
    return REF.sub(r'\1#', fml)

lines=['# XLSX metadata map — контроль / защита / аномалии','']
manifest=[]
for f in files:
    rel=os.path.basename(f); lines.append(f'## {rel}'); fe={'file':rel,'sheets':[]}
    try: wb=load_workbook(f, data_only=False)
    except Exception as e: lines.append(f'  ОШИБКА: {e}'); fe['error']=str(e); manifest.append(fe); continue
    try:
        names=wb.defined_names
        items=names.items() if hasattr(names,'items') else [(d.name,d) for d in names.definedName]
        dn={}
        for nm,d in items:
            try: dn[nm]=str(getattr(d,'value',getattr(d,'attr_text','')))[:80]
            except Exception: dn[nm]=''
        if dn: lines.append(f'- именованные диапазоны ({len(dn)}): '+', '.join(f'{k}={v}' for k,v in list(dn.items())[:20]))
        fe['defined_names']=dn
    except Exception: pass
    try:
        ext=[]
        for el in (getattr(wb,'_external_links',None) or []):
            try: ext.append(str(getattr(el.file_link,'Target','')))
            except Exception: pass
        if ext: lines.append(f'- ВНЕШНИЕ ССЫЛКИ ({len(ext)}): '+', '.join(ext[:10])); fe['external_links']=ext
    except Exception: pass
    if expect:
        missing=sorted(expect-set(wb.sheetnames))
        if missing: lines.append(f'- ⚠ ОТСУТСТВУЮТ ЛИСТЫ: {", ".join(missing)}'); fe['missing_sheets']=missing
    for ws in wb.worksheets:
        se={'sheet':ws.title,'dims':ws.dimensions}; seg=[f'### лист «{ws.title}» [{ws.dimensions}]']
        try: dvs=list(ws.data_validations.dataValidation)
        except Exception: dvs=[]
        if dvs:
            seg.append(f'  data validation ({len(dvs)}):'); se['data_validations']=[]
            for dv in dvs[:30]:
                f1=str(dv.formula1 or '')[:200].replace('\n',' ')
                seg.append(f'    - тип={dv.type} диапазон={dv.sqref} допустимо="{f1[:160]}" пусто={dv.allow_blank}')
                se['data_validations'].append({'type':dv.type,'sqref':str(dv.sqref),'formula1':f1,'allow_blank':dv.allow_blank})
        if val_only:
            if dvs: lines.append('\n'.join(seg))
            fe['sheets'].append(se); continue
        try: cfs=list(ws.conditional_formatting)
        except Exception: cfs=[]
        if cfs:
            seg.append(f'  условное форматирование ({len(cfs)}):'); se['cond_formatting']=[]
            for cf in cfs[:20]:
                try:
                    rtypes=[f'{r.type}:{(str(r.formula[0])[:50] if getattr(r,"formula",None) else (r.operator or ""))}' for r in cf.rules]
                    seg.append(f'    - {cf.sqref}: {"; ".join(rtypes[:6])}'); se['cond_formatting'].append({'sqref':str(cf.sqref),'rules':rtypes[:8]})
                except Exception as e:
                    seg.append(f'    - {getattr(cf,"sqref","?")}: <{e}>')
        try:
            if ws.protection and ws.protection.sheet: seg.append('  защита: лист защищён'); se['protected']=True
        except Exception: pass
        try:
            mr=[str(r) for r in ws.merged_cells.ranges]
            if mr: seg.append(f'  объединённых: {len(mr)} (напр. {", ".join(mr[:5])})'); se['merged']=len(mr)
        except Exception: pass
        try:
            hr=sum(1 for d in ws.row_dimensions.values() if d.hidden)
            hc=[k for k,d in ws.column_dimensions.items() if d.hidden]
            if hr or hc:
                seg.append(f'  скрыто: строк={hr} столбцов={len(hc)}'+((' ['+",".join(hc[:10])+']') if hc else ''))
                se['hidden_rows']=hr; se['hidden_cols']=hc
        except Exception: pass
        try:
            if ws.auto_filter and ws.auto_filter.ref: seg.append(f'  автофильтр: {ws.auto_filter.ref}'); se['autofilter']=str(ws.auto_filter.ref)
        except Exception: pass
        comments=[]; fmt_by_col=defaultdict(Counter); tmpl_by_col=defaultdict(Counter); tmpl_ex=defaultdict(dict)
        try:
            for row in ws.iter_rows():
                for c in row:
                    v=c.value
                    if c.comment is not None:
                        comments.append((c.coordinate,(c.comment.text or '')[:80].replace('\n',' ')))
                    if isinstance(v,str) and v.startswith('='):
                        t=templ(v); col=c.column_letter; tmpl_by_col[col][t]+=1
                        if t not in tmpl_ex[col]: tmpl_ex[col][t]=(c.coordinate,v[:80])
                    nf=getattr(c,'number_format',None)
                    if nf and nf!='General' and v not in (None,''): fmt_by_col[c.column_letter][nf]+=1
        except Exception as e:
            seg.append(f'  [pass error: {e}]')
        if comments:
            seg.append(f'  комментарии ({len(comments)}): '+'; '.join(f'{a}:{t}' for a,t in comments[:6]))
            se['comments']=[{'cell':a,'text':t} for a,t in comments[:30]]
        fmt_notable={}
        for col,cc in fmt_by_col.items():
            top=cc.most_common(1)[0][0]
            if '%' in top or '₽' in top or 'руб' in top.lower() or '0.0' in top: fmt_notable[col]=top
        if fmt_notable:
            seg.append('  форматы(%/₽): '+', '.join(f'{k}={v}' for k,v in list(fmt_notable.items())[:12])); se['number_formats']=fmt_notable
        anomalies=[]
        for col,cc in tmpl_by_col.items():
            total=sum(cc.values())
            if total<5: continue
            dom,domn=cc.most_common(1)[0]
            if domn/total>=0.6:
                for t,n in cc.items():
                    if t!=dom:
                        ex=tmpl_ex[col].get(t,('',''))
                        anomalies.append({'col':col,'odd':t[:50],'count':n,'example':ex[0],'dominant':dom[:50]})
        if anomalies:
            seg.append(f'  ⚠ аномалии формул ({len(anomalies)}): '+'; '.join(f'{a["col"]}@{a["example"]}≠эталон×{a["count"]}' for a in anomalies[:8]))
            se['formula_anomalies']=anomalies[:40]
        if len(seg)>1: lines.append('\n'.join(seg))
        fe['sheets'].append(se)
    lines.append(''); manifest.append(fe)
    try: wb.close()
    except Exception: pass

out_text='\n'.join(lines)
if out_path: Path(out_path).write_text(out_text,encoding='utf-8'); print(f'[OK] md {out_path} ({len(out_text)} chars)')
if out_json: Path(out_json).write_text(json.dumps(manifest,ensure_ascii=False,indent=1),encoding='utf-8'); print(f'[OK] json {out_json}')
if not out_path and not out_json: print(out_text[:6000])
